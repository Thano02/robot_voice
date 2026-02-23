# excel_handler.py
# Gestion de la lecture et écriture dans le fichier prospects.xlsx

import openpyxl
from datetime import datetime

FICHIER_EXCEL = "prospects.xlsx"

# Correspondance entre les colonnes Excel et les noms de champs
COLONNES = {
    "telephone":          1,   # A
    "prenom":             2,   # B
    "nom":                3,   # C
    "statut":             4,   # D
    "proprietaire":       5,   # E
    "type_logement":      6,   # F
    "annee_construction": 7,   # G
    "chauffage_actuel":   8,   # H
    "revenus":            9,   # I
    "eligibilite":        10,  # J
    "notes":              11,  # K
    "date_appel":         12,  # L
}


def lire_prospects():
    """
    Ouvre prospects.xlsx et retourne la liste des prospects
    ayant le statut 'À appeler', avec leur numéro de ligne Excel.
    """
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb.active

    prospects = []

    # Parcourt toutes les lignes en sautant l'en-tête (ligne 1)
    for numero_ligne, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        telephone   = row[0]  # Colonne A : Téléphone
        prenom      = row[1]  # Colonne B : Prénom
        nom_famille = row[2]  # Colonne C : Nom
        statut      = row[3]  # Colonne D : Statut

        # On ne traite que les prospects à appeler (accepte avec ou sans accent)
        if statut and str(statut).strip().lower() in ("à appeler", "a appeler"):
            # Formatte le numéro en E.164 (+33XXXXXXXXX)
            tel = str(telephone).strip() if telephone else ""
            if tel and not tel.startswith("+"):
                # Supprime le 0 initial si présent puis ajoute +33
                tel = "+33" + (tel[1:] if tel.startswith("0") else tel)

            nom_complet = " ".join(filter(None, [
                str(prenom).strip() if prenom else "",
                str(nom_famille).strip() if nom_famille else "",
            ])) or "Inconnu"

            prospects.append({
                "numero_ligne": numero_ligne,
                "telephone":    tel,
                "nom":          nom_complet,
            })

    wb.close()
    return prospects


def ecrire_resultat(numero_ligne: int, donnees: dict):
    """
    Écrit les résultats d'un appel dans la ligne Excel correspondante.

    :param numero_ligne: numéro de ligne dans Excel (commence à 2)
    :param donnees: dictionnaire avec les clés de COLONNES
    """
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb.active

    # Écrit chaque champ dans sa colonne
    for champ, valeur in donnees.items():
        if champ in COLONNES:
            col = COLONNES[champ]
            ws.cell(row=numero_ligne, column=col, value=valeur)

    # Horodatage de l'appel
    ws.cell(
        row=numero_ligne,
        column=COLONNES["date_appel"],
        value=datetime.now().strftime("%d/%m/%Y %H:%M")
    )

    wb.save(FICHIER_EXCEL)
    wb.close()
    print(f"[Excel] Ligne {numero_ligne} mise à jour.")


def marquer_pas_repondu(numero_ligne: int):
    """Marque le prospect comme 'Pas répondu'."""
    ecrire_resultat(numero_ligne, {"statut": "Pas répondu"})


def marquer_messagerie(numero_ligne: int):
    """Marque le prospect comme 'Messagerie'."""
    ecrire_resultat(numero_ligne, {"statut": "Messagerie"})


def creer_fichier_exemple():
    """
    Crée un fichier prospects.xlsx avec des données de test
    si le fichier n'existe pas encore.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # En-têtes
    entetes = [
        "telephone", "nom", "statut", "proprietaire",
        "type_logement", "annee_construction", "chauffage_actuel",
        "revenus", "eligibilite", "notes", "date_appel"
    ]
    ws.append(entetes)

    # Quelques prospects de test
    prospects_test = [
        ["+33612345678", "Jean Dupont",    "a appeler", "", "", "", "", "", "", "", ""],
        ["+33623456789", "Marie Martin",   "a appeler", "", "", "", "", "", "", "", ""],
        ["+33634567890", "Paul Bernard",   "Traite",    "", "", "", "", "", "", "", ""],
    ]
    for p in prospects_test:
        ws.append(p)

    wb.save(FICHIER_EXCEL)
    print(f"[Excel] Fichier '{FICHIER_EXCEL}' créé avec {len(prospects_test)} prospects de test.")


# --- Test rapide ---
if __name__ == "__main__":
    import os
    if not os.path.exists(FICHIER_EXCEL):
        creer_fichier_exemple()

    prospects = lire_prospects()
    print(f"\n{len(prospects)} prospect(s) à appeler :")
    for p in prospects:
        print(f"  Ligne {p['numero_ligne']} | {p['nom']} | {p['telephone']}")
